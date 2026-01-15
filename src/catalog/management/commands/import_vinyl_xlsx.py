# src/catalog/management/commands/import_vinyl_xlsx.py
from __future__ import annotations

from dataclasses import dataclass
from typing import Any

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction

from catalog.models import Artist, ArtistType, MediaItem, MediaType, SortBucket, CollectionOwner


# -----------------------------
# Legacy Access mappings
# -----------------------------
LEGACY_SORTKEY2_TO_BUCKET_CODE: dict[str, str | None] = {
    "1": "COUNTRY_AMERICANA",
    "2": "POP",
    "3": "ROCK",
    "4": "HARD_ROCK",
    "5": "RB_HIPHOP",
    "6": "BLUES_JAZZ",
    "7": "ALT_GRUNGE",
    "8": "SOUNDTRACKS",
    "9": "COMPS",
    "10": "HOLIDAY",
    "11": "NEWWAVE_SYNTH",
    "12": "MISC",
    "0": None,  # treat as invalid/unset (you can change this to "MISC" if you prefer)
}

LEGACY_SORTKEY3_TO_MEDIATYPE_NAME: dict[str, str | None] = {
    "10": "Standard LP",
    "11": "Valuable, Sealed, Special",  # Access label was shorter
    "14": "Premium Pressings",
    "15": "Box Set",
    "17": '7" Vinyl',
    "20": "Cassette Tape",
    "21": "CD",
    "0": None,  # treat as invalid/unset
}


# -----------------------------
# Helpers
# -----------------------------
def _clean_str(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _clean_int(v: Any) -> int | None:
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    try:
        # handles ints stored as floats in Excel
        return int(float(s))
    except Exception:
        return None


def _clean_bool(v: Any) -> bool | None:
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ("y", "yes", "true", "1"):
        return True
    if s in ("n", "no", "false", "0"):
        return False
    return None


def _build_header_map(ws) -> dict[str, int]:
    """Return {header_name: index} using the first row as headers."""
    headers = [(_clean_str(c.value)) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    return {h: i for i, h in enumerate(headers) if h}


def _first_present(col_map: dict[str, int], *names: str) -> int | None:
    for n in names:
        if n in col_map:
            return col_map[n]
    return None


@dataclass
class ImportStats:
    rows_seen: int = 0
    created_artists: int = 0
    created_items: int = 0
    updated_items: int = 0
    skipped_items: int = 0


class Command(BaseCommand):
    help = "Import vinyl rows from an Excel file (idempotent by MediaItem.master_key)."

    def add_arguments(self, parser):
        parser.add_argument("path", type=str, help="Path to Vinyl.xlsx (inside the container, e.g. /app/Vinyl.xlsx)")
        parser.add_argument("--sheet", type=str, default="Sheet1", help="Worksheet name (default: Sheet1)")
        parser.add_argument("--dry-run", action="store_true", help="Validate and simulate, but rollback all DB writes")
        parser.add_argument("--limit", type=int, default=None, help="Only process the first N data rows (after header)")
        parser.add_argument("--verbose", action="store_true", help="Print per-row actions")

    def handle(self, *args, **opts):
        path: str = opts["path"]
        sheet_name: str = opts["sheet"]
        dry_run: bool = opts["dry_run"]
        limit: int | None = opts["limit"]
        verbose: bool = opts["verbose"]

        wb = openpyxl.load_workbook(path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise SystemExit(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

        ws = wb[sheet_name]
        col_map = _build_header_map(ws)

        # ---- Header aliases (so you’re not hostage to exact names) ----
        c_master = _first_present(col_map, "MasterKey", "master_key", "MASTERKEY")
        c_artist_primary = _first_present(col_map, "ArtistPrimary", "artist_name_primary", "Artist Name Primary")
        c_artist_secondary = _first_present(col_map, "ArtistSecondary", "artist_name_secondary", "Artist Name Secondary")
        c_suffix = _first_present(col_map, "NameSuffix", "Suffix", "name_suffix")
        c_artist_type = _first_present(col_map, "ArtistType", "artist_type")
        c_title = _first_present(col_map, "AlbumTitle", "Title", "album_title")
        c_release_year = _first_present(col_map, "ReleaseYear", "release_year")
        c_sortkey2 = _first_present(col_map, "SortKey2", "sortkey2", "Bucket", "SortBucket")
        c_sortkey3 = _first_present(col_map, "SortKey3", "sortkey3", "MediaType")
        c_special = _first_present(col_map, "Special", "special", "Owned", "IsOwned")

        missing = []
        if c_master is None: missing.append("MasterKey")
        if c_artist_primary is None: missing.append("ArtistPrimary")
        if c_artist_secondary is None: missing.append("ArtistSecondary")
        if c_artist_type is None: missing.append("ArtistType")
        if c_title is None: missing.append("AlbumTitle")
        if c_release_year is None: missing.append("ReleaseYear")
        if c_sortkey2 is None: missing.append("SortKey2")
        if c_sortkey3 is None: missing.append("SortKey3")
        if c_special is None: missing.append("Special")

        if missing:
            raise SystemExit(
                f"Missing required columns: {missing}\n"
                f"Found headers: {list(col_map.keys())}"
            )

        # Preload lookup tables once
        buckets_by_code = {b.code: b for b in SortBucket.objects.all()}
        media_types_by_name = {m.name: m for m in MediaType.objects.select_related("default_zone").all()}

        stats = ImportStats()

        # True dry-run: write inside a transaction and force rollback at the end
        with transaction.atomic():
            for r_index, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if limit is not None and stats.rows_seen >= limit:
                    break

                stats.rows_seen += 1

                master_key = _clean_str(row[c_master].value)
                artist_primary = _clean_str(row[c_artist_primary].value)
                artist_secondary = _clean_str(row[c_artist_secondary].value)
                name_suffix = _clean_str(row[c_suffix].value) if c_suffix is not None else ""
                artist_type_raw = _clean_str(row[c_artist_type].value).upper()
                title = _clean_str(row[c_title].value)
                release_year = _clean_int(row[c_release_year].value)
                sortkey2_raw = _clean_str(row[c_sortkey2].value)
                sortkey3_raw = _clean_str(row[c_sortkey3].value)
                special_raw = row[c_special].value
                special = _clean_bool(special_raw)

                # Basic row validation
                if not master_key or not title or not artist_primary:
                    stats.skipped_items += 1
                    if verbose:
                        self.stdout.write(f"Row {r_index}: SKIP (missing master_key/title/artist_primary)")
                    continue

                # ---- ArtistType ----
                if artist_type_raw not in ("PERSON", "BAND"):
                    artist_type = ArtistType.BAND
                else:
                    artist_type = artist_type_raw

                # ---- SortKey2 -> bucket ----
                bucket_code = LEGACY_SORTKEY2_TO_BUCKET_CODE.get(sortkey2_raw, sortkey2_raw)

                if bucket_code is None:
                    raise SystemExit(f"Row {r_index}: SortKey2 '{sortkey2_raw}' is unmapped/invalid")

                bucket = buckets_by_code.get(bucket_code)
                if bucket is None:
                    # allow if sheet already contains codes/names, just in case
                    bucket = SortBucket.objects.filter(code__iexact=bucket_code).first() or SortBucket.objects.filter(
                        name__iexact=bucket_code
                    ).first()

                if bucket is None:
                    raise SystemExit(
                        f"Row {r_index}: Unknown SortKey2 '{sortkey2_raw}' -> '{bucket_code}' (not seeded?)"
                    )

                # ---- SortKey3 -> media_type ----
                media_name = LEGACY_SORTKEY3_TO_MEDIATYPE_NAME.get(sortkey3_raw, sortkey3_raw)

                if media_name is None:
                    raise SystemExit(f"Row {r_index}: SortKey3 '{sortkey3_raw}' is unmapped/invalid")

                media_type = media_types_by_name.get(media_name)
                if media_type is None:
                    media_type = MediaType.objects.filter(name__iexact=media_name).select_related("default_zone").first()

                if media_type is None:
                    raise SystemExit(
                        f"Row {r_index}: Unknown media type '{media_name}' derived from SortKey3 '{sortkey3_raw}'"
                    )

                # ---- Owner mapping ----
                # Your current logic: Special==True means "BIL", otherwise "ME"
                # If you later want Special to mean "owned but special", we’ll change this.
                owner = CollectionOwner.BIL if special else CollectionOwner.ME

                # ---- Upsert Artist ----
                artist_defaults = {
                    "artist_type": artist_type,
                    "artist_name_primary": artist_primary,
                    "artist_name_secondary": artist_secondary if artist_type == ArtistType.PERSON else "",
                    "name_suffix": name_suffix if artist_type == ArtistType.PERSON else "",
                }

                if artist_type == ArtistType.PERSON:
                    if not artist_secondary.strip():
                        raise SystemExit(f"Row {r_index}: PERSON artist missing last name (ArtistSecondary)")
                    artist, created_artist = Artist.objects.get_or_create(
                        artist_type=ArtistType.PERSON,
                        artist_name_primary=artist_primary,
                        artist_name_secondary=artist_secondary,
                        defaults=artist_defaults,
                    )
                else:
                    artist, created_artist = Artist.objects.get_or_create(
                        artist_type=ArtistType.BAND,
                        artist_name_primary=artist_primary,
                        defaults=artist_defaults,
                    )

                if created_artist:
                    stats.created_artists += 1

                # ---- Upsert MediaItem by master_key ----
                item_defaults = {
                    "artist": artist,
                    "title": title,
                    "release_year": release_year,
                    "pressing_year": release_year,  # per your current plan (can diverge later)
                    "media_type": media_type,
                    "bucket": bucket,
                    "owner": owner,
                }

                obj, created_item = MediaItem.objects.update_or_create(
                    master_key=master_key,
                    defaults=item_defaults,
                )

                if created_item:
                    stats.created_items += 1
                    action = "CREATE"
                else:
                    stats.updated_items += 1
                    action = "UPDATE"

                if verbose:
                    self.stdout.write(f"Row {r_index}: {action} {master_key} | {artist.display_name} — {title}")

            if dry_run:
                # Force rollback of the atomic transaction
                transaction.set_rollback(True)

        # ---- Summary ----
        self.stdout.write("")
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN complete (all changes rolled back)."))
        else:
            self.stdout.write(self.style.SUCCESS("Import complete."))

        self.stdout.write(f"Rows seen:       {stats.rows_seen}")
        self.stdout.write(f"Artists created: {stats.created_artists}")
        self.stdout.write(f"Items created:   {stats.created_items}")
        self.stdout.write(f"Items updated:   {stats.updated_items}")
        self.stdout.write(f"Items skipped:   {stats.skipped_items}")
